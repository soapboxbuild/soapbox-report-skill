#!/usr/bin/env python3
"""
test_build_xlsx.py — Smoke test for build_xlsx.py using Prose Frontier sample data.

Run:
    python scripts/test_build_xlsx.py

Exits 0 on success, 1 on failure.
"""

import json
import os
import sys
import tempfile
from typing import Any

# Allow running from repo root or from scripts/
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from build_xlsx import build_workbook

# ---------------------------------------------------------------------------
# Sample data — Prose Frontier (Prosper, TX multifamily)
# Conforms to templates/rsra/schema.json
# ---------------------------------------------------------------------------

PROSE_FRONTIER_DATA = {
    "property": {
        "name": "Prose Frontier",
        "address": "4400 Prairie Crossing, Prosper, TX 75078",
        "type": "multifamily",
        "units": 324,
        "gfa_sqft": 311922,
        "year_built": 2023,
        "zip": "75078",
    },
    "deal_signal": {
        "level": "Low Risk",
        "narrative": (
            "New all-electric Class A construction in a deregulated Texas market — "
            "no active BPS exposure, low transition risk over the hold period. "
            "Modest CapEx focused on smart-home compliance and EV readiness is "
            "accretive to NOI and exit cap rate."
        ),
    },
    "decarb_plan": [
        {
            "measure": "ENERGY STAR appliance package — replace iApts non-compliant units",
            "timing": "Early Yr1",
            "capex_per_unit": 250,
            "capex_total": 81000,
            "incentive_program": "§45L tax credit — $2,500/unit for ENERGY STAR multifamily",
            "financial_impact_type": "§45L tax credit",
            "financial_impact_timing": "Year 1 one-time",
            "financial_impact_value": "$810,000 (one-time, 10× CapEx)",
        },
        {
            "measure": "EV charging infrastructure — 10% of spaces, DCFC-ready conduit",
            "timing": "Early Yr2",
            "capex_per_unit": 185,
            "capex_total": 59940,
            "incentive_program": "NEVI / IRA §30C — 30% tax credit on EV charging equipment",
            "financial_impact_type": "ITC tax credit",
            "financial_impact_timing": "Year 2 one-time",
            "financial_impact_value": "~$17,982 (30% of CapEx)",
        },
        {
            "measure": "Whole-building submetering — Conservice manual meter program",
            "timing": "Early Yr1",
            "capex_per_unit": 140,
            "capex_total": 45360,
            "incentive_program": None,
            "financial_impact_type": "Common area expense reduction",
            "financial_impact_timing": "Annual ongoing",
            "financial_impact_value": "$14,400/yr (est.)",
        },
        {
            "measure": "Rooftop solar feasibility study + interconnection application",
            "timing": "Mid Yr3",
            "capex_per_unit": 300,
            "capex_total": 97200,
            "incentive_program": "ITC §48E — 30% solar tax credit",
            "financial_impact_type": "ITC tax credit",
            "financial_impact_timing": "Year 3 one-time",
            "financial_impact_value": "~$29,160 (30% of CapEx)",
        },
    ],
    # decarb_plan_total intentionally omitted — must be auto-computed
    "emissions_profile": {
        "fuel_profile": "~100% electric (confirmed, ESG DD report June 2023)",
        "utility_structure": (
            "Resident electric submetered via iApts; landlord pays common area electric "
            "and water. No gas service to the property."
        ),
        "baseline_emissions": "~42 kgCO₂e/m²yr (est., all-electric multifamily TX grid benchmark)",
        "crrem_pathway": (
            "~8% above 2030 CRREM 1.5°C target for US multifamily. "
            "Planned ENERGY STAR + submetering measures close the gap by Yr2."
        ),
        "regulation": (
            "Low — no active BPS in Prosper, TX. Monitor Colorado SB 19-096 if "
            "fund expands to CO. TX deregulated grid; no mandatory benchmarking."
        ),
        "eui_kbtu": 52.4,
        "energy_star_score": None,  # not yet enrolled
    },
    "seller_questions": [
        "Can you provide the last 12 months of whole-building utility bills "
        "or Conservice sub-meter data to confirm the all-electric fuel profile?",
        "Has the iApts technology package been evaluated against ENERGY STAR "
        "multifamily certification requirements? Are any units currently certified?",
        "Are there any pending or anticipated utility infrastructure charges "
        "related to the DCFC-ready conduit or EV parking buildout?",
    ],
    "data_quality": "Medium — confirmed fuel profile; CapEx benchmarked",
    "prepared_by": "Aris / Audette Sustainability Intelligence",
    "prepared_for": "Stoneweg US — Acquisitions",
    "report_date": "2026-06-16",
}

STONEWEG_BRAND = {
    "name": "Stoneweg US",
    "primary_color": "#1B2A3B",
    "secondary_color": "#B08D57",
    "accent_color": "#F5F3F0",
    "highlight_color": "#3D7A5C",
    "text_color": "#1B2A3B",
    "text_muted": "#6B7280",
    "border_color": "#E5E2DD",
    "font_heading": "Inter",
    "font_body": "Inter",
    "font_mono": "JetBrains Mono",
}

TEMPLATES_DIR = os.path.join(
    os.path.dirname(os.path.abspath(__file__)), "..", "templates"
)


# ---------------------------------------------------------------------------
# Test helpers
# ---------------------------------------------------------------------------

def assert_equal(label: str, actual: Any, expected: Any) -> None:
    if actual != expected:
        print(f"  FAIL  {label}: expected {expected!r}, got {actual!r}")
        return False
    print(f"  PASS  {label}")
    return True


def assert_true(label: str, condition: bool) -> bool:
    if not condition:
        print(f"  FAIL  {label}")
        return False
    print(f"  PASS  {label}")
    return True


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------

def test_rsra_workbook() -> bool:
    """Build RSRA workbook from Prose Frontier data and verify sheet structure."""
    import openpyxl

    print("\n[test_rsra_workbook]")
    failures = 0

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        output_path = tmp.name

    try:
        build_workbook(
            template="rsra",
            data=PROSE_FRONTIER_DATA.copy(),
            brand=STONEWEG_BRAND,
            output_path=output_path,
            templates_dir=TEMPLATES_DIR,
        )

        wb = openpyxl.load_workbook(output_path)
        sheet_names = wb.sheetnames
        print(f"  Sheets built: {sheet_names}")

        # 1. Expected sheets present
        expected_sheets = [
            "RSRA Summary",
            "Decarb Plan",
            "Emissions Profile",
            "Seller Questions",
            "Incentives",
        ]
        for name in expected_sheets:
            if not assert_true(f"sheet '{name}' exists", name in sheet_names):
                failures += 1

        # 2. RSRA Summary is first
        if not assert_true("RSRA Summary is first sheet", sheet_names[0] == "RSRA Summary"):
            failures += 1

        # 3. decarb_plan_total was auto-computed (Total CapEx row has a value)
        ws_summary = wb["RSRA Summary"]
        found_capex = False
        for row in ws_summary.iter_rows(values_only=True):
            if row[0] == "Total CapEx" and row[1] is not None:
                found_capex = True
                expected_total = 81000 + 59940 + 45360 + 97200  # 283500
                if not assert_true(
                    f"Total CapEx = {expected_total}",
                    abs(float(row[1]) - expected_total) < 0.01,
                ):
                    failures += 1
                break
        if not assert_true("Total CapEx row found in RSRA Summary", found_capex):
            failures += 1

        # 4. Decarb Plan has header row + 4 data rows + 1 totals row = ≥5 rows
        ws_decarb = wb["Decarb Plan"]
        row_count = ws_decarb.max_row
        if not assert_true(f"Decarb Plan has ≥5 rows (has {row_count})", row_count >= 5):
            failures += 1

        # 5. Seller Questions sheet is present and has data rows
        ws_sq = wb["Seller Questions"]
        sq_rows = ws_sq.max_row
        if not assert_true(f"Seller Questions has ≥2 rows (has {sq_rows})", sq_rows >= 2):
            failures += 1

        # 6. Incentives sheet filters correctly — only rows with incentive_program
        ws_inc = wb["Incentives"]
        inc_rows = ws_inc.max_row
        # 3 of 4 decarb rows have incentive_program (one is None)
        if not assert_true(
            f"Incentives sheet has ≥3 data rows (has {inc_rows - 1})", inc_rows >= 4
        ):
            failures += 1

        # 7. Emissions Profile sheet exists and has data
        ws_ep = wb["Emissions Profile"]
        ep_rows = ws_ep.max_row
        if not assert_true(f"Emissions Profile has ≥3 rows (has {ep_rows})", ep_rows >= 3):
            failures += 1

        # 8. Total CapEx cell on the RSRA Summary sheet should carry a currency format
        for row in ws_summary.iter_rows():
            if row[0].value == "Total CapEx" and row[1].value is not None:
                fmt = row[1].number_format or ""
                if not assert_true(
                    f"Total CapEx cell has currency format (got '{fmt}')",
                    "$" in fmt,
                ):
                    failures += 1
                break

    finally:
        try:
            os.unlink(output_path)
        except OSError:
            pass

    return failures == 0


def test_auto_derive_fallback() -> bool:
    """When no xlsx.json exists for a template, sheets are auto-derived from arrays in data."""
    print("\n[test_auto_derive_fallback]")
    import openpyxl

    failures = 0
    fake_data = {
        "report_date": "2026-06-16",
        "items": [
            {"name": "Alpha", "value": 1},
            {"name": "Beta", "value": 2},
        ],
    }

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        output_path = tmp.name

    try:
        build_workbook(
            template="_nonexistent_template_",
            data=fake_data,
            brand=STONEWEG_BRAND,
            output_path=output_path,
            templates_dir=TEMPLATES_DIR,
        )
        wb = openpyxl.load_workbook(output_path)
        names = wb.sheetnames
        print(f"  Sheets: {names}")

        if not assert_true("Summary sheet exists (auto-generated)", "Summary" in names):
            failures += 1
        if not assert_true("Items sheet exists (auto-derived)", "Items" in names):
            failures += 1
        if not assert_true("Summary is first", names[0] == "Summary"):
            failures += 1

    finally:
        try:
            os.unlink(output_path)
        except OSError:
            pass

    return failures == 0


def test_omit_if_empty_seller_questions() -> bool:
    """When seller_questions is absent, Seller Questions sheet must be omitted."""
    print("\n[test_omit_if_empty_seller_questions]")
    import openpyxl

    data = PROSE_FRONTIER_DATA.copy()
    data.pop("seller_questions", None)

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        output_path = tmp.name

    failures = 0
    try:
        build_workbook(
            template="rsra",
            data=data,
            brand=STONEWEG_BRAND,
            output_path=output_path,
            templates_dir=TEMPLATES_DIR,
        )
        wb = openpyxl.load_workbook(output_path)
        names = wb.sheetnames
        print(f"  Sheets: {names}")

        if not assert_true(
            "Seller Questions omitted when data absent", "Seller Questions" not in names
        ):
            failures += 1

    finally:
        try:
            os.unlink(output_path)
        except OSError:
            pass

    return failures == 0


def test_cli_interface() -> bool:
    """CLI: --template + --data (JSON string) + --brand (JSON string) + --output."""
    print("\n[test_cli_interface]")
    import subprocess

    failures = 0
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        output_path = tmp.name

    script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "build_xlsx.py")
    templates_dir = os.path.abspath(TEMPLATES_DIR)

    cmd = [
        sys.executable, script,
        "--template", "rsra",
        "--data", json.dumps(PROSE_FRONTIER_DATA),
        "--brand", json.dumps(STONEWEG_BRAND),
        "--output", output_path,
        "--templates-dir", templates_dir,
    ]

    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=60)
        if not assert_true(
            f"CLI exits 0 (got {result.returncode})", result.returncode == 0
        ):
            print(f"    stderr: {result.stderr[:400]}")
            failures += 1
        if not assert_true(
            "CLI prints output path to stdout",
            output_path in result.stdout.strip(),
        ):
            print(f"    stdout: {result.stdout!r}")
            failures += 1
        if not assert_true("Output file was created", os.path.exists(output_path)):
            failures += 1
        if not assert_true("No chart build warnings in stderr", "WARNING" not in result.stderr):
            print(f"    stderr: {result.stderr[:400]}")
            failures += 1
    finally:
        try:
            os.unlink(output_path)
        except OSError:
            pass

    return failures == 0


def test_cli_config_file() -> bool:
    """CLI: --config JSON file with all params."""
    print("\n[test_cli_config_file]")
    import subprocess

    failures = 0
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as xlsx_tmp:
        output_path = xlsx_tmp.name
    with tempfile.NamedTemporaryFile(
        mode="w", suffix=".json", delete=False, encoding="utf-8"
    ) as cfg_tmp:
        json.dump(
            {
                "template": "rsra",
                "data": PROSE_FRONTIER_DATA,
                "brand": STONEWEG_BRAND,
                "output": output_path,
                "templates_dir": os.path.abspath(TEMPLATES_DIR),
            },
            cfg_tmp,
        )
        cfg_path = cfg_tmp.name

    script = os.path.join(os.path.dirname(os.path.abspath(__file__)), "build_xlsx.py")

    try:
        result = subprocess.run(
            [sys.executable, script, "--config", cfg_path],
            capture_output=True, text=True, timeout=60,
        )
        if not assert_true(f"CLI --config exits 0 (got {result.returncode})", result.returncode == 0):
            print(f"    stderr: {result.stderr[:400]}")
            failures += 1
        if not assert_true("Output file created via --config", os.path.exists(output_path)):
            failures += 1
    finally:
        for p in [output_path, cfg_path]:
            try:
                os.unlink(p)
            except OSError:
                pass

    return failures == 0


# ---------------------------------------------------------------------------
# Runner
# ---------------------------------------------------------------------------

def main() -> None:
    tests = [
        test_rsra_workbook,
        test_omit_if_empty_seller_questions,
        test_auto_derive_fallback,
        test_cli_interface,
        test_cli_config_file,
    ]

    passed = 0
    failed = 0
    for test_fn in tests:
        try:
            ok = test_fn()
        except Exception as exc:
            import traceback
            print(f"  ERROR in {test_fn.__name__}: {exc}")
            traceback.print_exc()
            ok = False
        if ok:
            passed += 1
        else:
            failed += 1

    print(f"\n{'=' * 50}")
    print(f"Results: {passed} passed, {failed} failed")
    if failed:
        sys.exit(1)
    print("All tests passed.")


if __name__ == "__main__":
    main()
