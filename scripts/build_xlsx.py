#!/usr/bin/env python3
"""
build_xlsx.py — RSRA Excel companion builder for soapbox-report-skill.

Called by the report-renderer agent after every render. Reads xlsx.json from the
template directory, applies brand colors, and emits a production-quality .xlsx.

Usage (inline JSON args):
    python build_xlsx.py \\
        --template rsra \\
        --data '{"property": {...}, ...}' \\
        --brand '{"primary_color": "#1B2A3B", ...}' \\
        --output /path/to/output.xlsx \\
        --templates-dir /path/to/soapbox-report-skill/templates

Usage (JSON config file):
    python build_xlsx.py --config /path/to/render_config.json

Where render_config.json keys: template, data, brand, output, templates_dir.
data and brand may be JSON strings or pre-parsed objects.
"""

import argparse
import json
import os
import sys
from collections import defaultdict
from typing import Any, Dict, List, Optional

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as e:
    print(f"ERROR: openpyxl is required. Run: pip install openpyxl  ({e})", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Colour helpers
# ---------------------------------------------------------------------------

# Maps semantic token names to brand dict keys
_SEMANTIC_TO_BRAND_KEY: Dict[str, str] = {
    "primary":    "primary_color",
    "secondary":  "secondary_color",
    "accent":     "accent_color",
    "highlight":  "highlight_color",
    "text":       "text_color",
    "text_muted": "text_muted",
}


def _strip_hash(hex_color: str) -> str:
    """Remove leading '#', return bare 6-char hex. E.g. '#1B2A3B' → '1B2A3B'."""
    return hex_color.lstrip("#").upper()


def hex_to_argb(hex_color: str) -> str:
    """Convert '#1B2A3B' or '1B2A3B' to 'FF1B2A3B' for openpyxl Font/Fill colors."""
    if not hex_color:
        return "FF000000"
    h = _strip_hash(hex_color)
    if len(h) == 6:
        return f"FF{h}"
    if len(h) == 8:
        return h
    return "FF000000"


def resolve_color(spec: Any, brand: dict) -> Optional[str]:
    """
    Resolve a color spec to a bare 6-char hex string (no '#', no alpha prefix).

    Handles three forms:
      "brand.primary_color"    → brand["primary_color"] → strip '#'
      "primary" / "accent" / … → SEMANTIC_TO_BRAND_KEY lookup → strip '#'
      "#D1FAE5" / "D1FAE5"    → strip '#' and return
      "none" / None / ""       → return None (no fill)
    """
    if spec is None or spec == "" or spec == "none":
        return None
    s = str(spec).strip()
    if s.startswith("brand."):
        key = s[len("brand."):]
        raw = brand.get(key, "")
        return _strip_hash(raw) if raw else None
    if s in _SEMANTIC_TO_BRAND_KEY:
        raw = brand.get(_SEMANTIC_TO_BRAND_KEY[s], "")
        return _strip_hash(raw) if raw else None
    if s.startswith("#") or (len(s) == 6 and all(c in "0123456789ABCDEFabcdef" for c in s)):
        return _strip_hash(s)
    return None


def make_fill(hex6: Optional[str]) -> Optional[PatternFill]:
    """Return a solid PatternFill from a bare-hex color, or None."""
    if not hex6:
        return None
    return PatternFill(fill_type="solid", fgColor=f"FF{hex6.upper()}")


def make_font(
    name: str = "Calibri",
    size: int = 9,
    bold: bool = False,
    italic: bool = False,
    color: Optional[str] = None,   # bare 6-hex or #hex
) -> Font:
    kwargs: Dict[str, Any] = {"name": name, "size": size, "bold": bold, "italic": italic}
    if color:
        kwargs["color"] = hex_to_argb(color)
    return Font(**kwargs)


def make_border(color_hex6: str = "E5E7EB") -> Border:
    side = Side(style="thin", color=f"FF{color_hex6.upper()}")
    return Border(bottom=side)


# ---------------------------------------------------------------------------
# Data path resolver
# ---------------------------------------------------------------------------

def resolve_path(data: Any, dotpath: str) -> Any:
    """Traverse nested dict by dot-notation path. Returns None on any miss."""
    if not dotpath:
        return None
    parts = dotpath.split(".")
    node = data
    for part in parts:
        if not isinstance(node, dict):
            return None
        node = node.get(part)
        if node is None:
            return None
    return node


# ---------------------------------------------------------------------------
# Sheet-name sanitizer
# ---------------------------------------------------------------------------

def sanitize_sheet_name(name: str) -> str:
    """Strip Excel-illegal chars and truncate to 31 characters."""
    import re
    name = re.sub(r'[:\\/?\*\[\]]', "", name)
    return name[:31]


# ---------------------------------------------------------------------------
# decarb_plan_total auto-compute
# ---------------------------------------------------------------------------

def ensure_decarb_plan_total(data: dict) -> None:
    """
    Compute decarb_plan_total from decarb_plan if the top-level key is absent or
    incomplete. Mutates data in-place — called once before any sheet is built.
    """
    rows: List[dict] = data.get("decarb_plan", [])
    if not rows:
        return
    existing = data.get("decarb_plan_total") or {}
    if "capex_total" not in existing:
        existing["capex_total"] = sum(r.get("capex_total", 0) or 0 for r in rows)
    if "capex_per_unit" not in existing:
        vals = [r["capex_per_unit"] for r in rows if r.get("capex_per_unit") is not None]
        if vals:
            existing["capex_per_unit"] = sum(vals)
    data["decarb_plan_total"] = existing


# ---------------------------------------------------------------------------
# key_value sheet builder
# ---------------------------------------------------------------------------

def build_key_value_sheet(
    ws, sheet_spec: dict, data: dict, brand: dict, wb_style: dict
) -> None:
    """Two-column label / value sheet."""
    rows_spec: List[dict] = sheet_spec.get("rows", [])
    columns_spec: List[dict] = sheet_spec.get("columns", [])
    style_spec: dict = sheet_spec.get("style", {})

    body_font_name: str = wb_style.get("body_font", brand.get("font_body", "Calibri"))
    body_font_size: int = wb_style.get("body_font_size", 9)
    header_font_size: int = wb_style.get("header_font_size", 9)
    header_fill_hex: Optional[str] = resolve_color(
        wb_style.get("header_fill_color", "brand.primary_color"), brand
    )
    header_font_hex: Optional[str] = resolve_color(
        wb_style.get("header_font_color", "#FFFFFF"), brand
    ) or "FFFFFF"
    accent_hex: Optional[str] = resolve_color("accent", brand) or resolve_color(
        brand.get("accent_color", "#F0F4F8"), brand
    )
    text_hex: str = _strip_hash(brand.get("text_color", "#0F1923"))
    border_hex: str = _strip_hash(brand.get("border_color", "#E5E7EB"))
    alt_fill_hex: Optional[str] = _strip_hash(
        wb_style.get("alternate_row_fill", "#F9FAFB") or "#F9FAFB"
    )

    signal_row_style: dict = style_spec.get("signal_level_row", {})
    signal_fills: dict = signal_row_style.get("fill_by_value", {})
    label_col_style: dict = style_spec.get("label_column", {})

    label_bold: bool = label_col_style.get("bold", True)
    label_fill_spec = label_col_style.get("fill", "accent")
    label_fill_hex: Optional[str] = resolve_color(label_fill_spec, brand) or accent_hex

    # Write column headers (row 1)
    headers = [c.get("header", "") for c in columns_spec]
    for ci, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = make_font(
            name=body_font_name,
            size=header_font_size,
            bold=True,
            color=header_font_hex,
        )
        if header_fill_hex:
            cell.fill = make_fill(header_fill_hex)
        cell.alignment = Alignment(wrap_text=False, vertical="center")

    # Set column widths from spec (never auto-size — spec widths are intentional)
    for ci, col_spec in enumerate(columns_spec, start=1):
        w = col_spec.get("width")
        if w:
            ws.column_dimensions[get_column_letter(ci)].width = float(w)

    ws.freeze_panes = "A2"

    data_row_idx = 0  # for alternating fills
    row_num = 2

    for row_spec in rows_spec:
        if row_spec.get("type") == "spacer":
            ws.row_dimensions[row_num].height = 8
            row_num += 1
            continue

        label: str = row_spec.get("label", "")
        dotpath: str = row_spec.get("data_path", "")
        value = resolve_path(data, dotpath) if dotpath else None

        if row_spec.get("omit_if_null") and (value is None or value == ""):
            continue

        row_style: str = row_spec.get("style", "")
        fmt: Optional[str] = row_spec.get("format")
        row_height: Optional[int] = row_spec.get("row_height")

        # Determine fill colors for this row
        row_label_fill_hex = label_fill_hex
        row_value_fill_hex = None
        row_bold_override = False

        if row_style == "signal_level":
            # Match by substring to handle enum variants with em-dashes
            matched: Optional[str] = None
            if value:
                for key_str, fill_val in signal_fills.items():
                    if key_str in str(value):
                        matched = _strip_hash(fill_val)
                        break
            if matched:
                row_label_fill_hex = matched
                row_value_fill_hex = matched
            row_bold_override = signal_row_style.get("bold", True)

        # Label cell
        label_cell = ws.cell(row=row_num, column=1, value=label)
        label_cell.font = make_font(
            name=body_font_name,
            size=body_font_size,
            bold=label_bold or row_bold_override,
            color=text_hex,
        )
        if row_label_fill_hex:
            label_cell.fill = make_fill(row_label_fill_hex)
        label_cell.alignment = Alignment(vertical="center", wrap_text=False)
        label_cell.border = make_border(border_hex)

        # Value cell
        value_cell = ws.cell(row=row_num, column=2, value=value)
        value_cell.font = make_font(
            name=body_font_name,
            size=body_font_size,
            bold=row_bold_override,
            color=text_hex,
        )
        if row_value_fill_hex:
            value_cell.fill = make_fill(row_value_fill_hex)
        elif data_row_idx % 2 == 1 and alt_fill_hex:
            value_cell.fill = make_fill(alt_fill_hex)

        wrap_value = row_style == "wrap" or (
            len(columns_spec) > 1 and columns_spec[1].get("wrap", False)
        )
        value_cell.alignment = Alignment(vertical="top", wrap_text=wrap_value)
        value_cell.border = make_border(border_hex)

        if fmt:
            value_cell.number_format = fmt

        if row_height:
            ws.row_dimensions[row_num].height = row_height

        row_num += 1
        data_row_idx += 1


# ---------------------------------------------------------------------------
# table sheet builder
# ---------------------------------------------------------------------------

def build_table_sheet(
    ws, sheet_spec: dict, data: dict, brand: dict, wb_style: dict
) -> None:
    """Tabular sheet with headers, data rows, optional totals row, charts, note."""
    dotpath: str = sheet_spec.get("data_path", "")
    columns_spec: List[dict] = sheet_spec.get("columns", [])
    totals_spec: Optional[dict] = sheet_spec.get("totals_row")
    cond_fmt: dict = sheet_spec.get("conditional_formatting", {})
    freeze: Optional[str] = sheet_spec.get("freeze_panes", "A2")
    note: Optional[str] = sheet_spec.get("note") or sheet_spec.get("footer_note")
    header_note: Optional[str] = sheet_spec.get("header_note")
    charts_spec: List[dict] = sheet_spec.get("charts", [])
    filter_spec: Optional[dict] = sheet_spec.get("filter")

    body_font_name: str = wb_style.get("body_font", brand.get("font_body", "Calibri"))
    body_font_size: int = wb_style.get("body_font_size", 9)
    header_font_size: int = wb_style.get("header_font_size", 9)
    header_fill_hex: Optional[str] = resolve_color(
        wb_style.get("header_fill_color", "brand.primary_color"), brand
    )
    header_font_hex: str = resolve_color(
        wb_style.get("header_font_color", "#FFFFFF"), brand
    ) or "FFFFFF"
    alt_fill_hex: Optional[str] = _strip_hash(
        wb_style.get("alternate_row_fill", "#F9FAFB") or "#F9FAFB"
    )
    border_hex: str = _strip_hash(brand.get("border_color", "#E5E7EB"))
    text_hex: str = _strip_hash(brand.get("text_color", "#0F1923"))
    text_muted_hex: str = _strip_hash(brand.get("text_muted", "#6B7280"))

    # Resolve data
    rows_data: List[Any] = resolve_path(data, dotpath) or []
    if not isinstance(rows_data, list):
        rows_data = []

    # Apply filter
    if filter_spec:
        field = filter_spec.get("field")
        if filter_spec.get("not_null") and field:
            rows_data = [
                r for r in rows_data
                if isinstance(r, dict) and r.get(field) not in (None, "", [])
            ]

    # Drop omit_if_all_null columns
    active_columns: List[dict] = []
    for col in columns_spec:
        key = col.get("data_key")
        if col.get("omit_if_all_null") and key and key != "_value":
            all_null = all(
                (isinstance(r, dict) and r.get(key) in (None, ""))
                for r in rows_data
            )
            if all_null:
                continue
        active_columns.append(col)

    # Header note row (above headers)
    start_row = 1
    if header_note:
        note_cell = ws.cell(row=1, column=1, value=header_note)
        note_cell.font = make_font(
            name=body_font_name, size=body_font_size - 1,
            italic=True, color=text_muted_hex,
        )
        note_cell.alignment = Alignment(wrap_text=True)
        if len(active_columns) > 1:
            ws.merge_cells(
                start_row=1, start_column=1,
                end_row=1, end_column=len(active_columns),
            )
        ws.row_dimensions[1].height = 30
        start_row = 2

    header_row = start_row

    # Header row
    for ci, col in enumerate(active_columns, start=1):
        cell = ws.cell(row=header_row, column=ci, value=col.get("header", ""))
        cell.font = make_font(
            name=body_font_name, size=header_font_size, bold=True, color=header_font_hex,
        )
        if header_fill_hex:
            cell.fill = make_fill(header_fill_hex)
        cell.alignment = Alignment(
            wrap_text=col.get("wrap", False), vertical="center"
        )
        # Use spec width; no auto-sizing for explicitly specified columns
        w = col.get("width")
        if w:
            ws.column_dimensions[get_column_letter(ci)].width = float(w)

    # Auto-filter on headers
    if active_columns:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(active_columns))}{header_row}"
        )

    if freeze:
        ws.freeze_panes = freeze

    # Data rows
    first_data_row = header_row + 1
    for ri, row in enumerate(rows_data, start=0):
        excel_row = first_data_row + ri
        use_alt = (ri % 2 == 1) and bool(alt_fill_hex)

        for ci, col in enumerate(active_columns, start=1):
            col_type = col.get("type")
            key = col.get("data_key", "")

            if col_type == "row_index":
                value = ri + 1
            elif key == "_value":
                value = row if not isinstance(row, dict) else str(row)
            else:
                value = row.get(key) if isinstance(row, dict) else None

            cell = ws.cell(row=excel_row, column=ci, value=value)
            cell.font = make_font(
                name=body_font_name, size=body_font_size, color=text_hex,
            )
            cell.alignment = Alignment(
                wrap_text=col.get("wrap", False), vertical="top",
            )
            cell.border = make_border(border_hex)

            if use_alt:
                cell.fill = make_fill(alt_fill_hex)

            # Conditional formatting by column value (exact-string match)
            if key in cond_fmt and isinstance(value, str):
                rule = cond_fmt[key].get(value)
                if rule and "fill" in rule:
                    cell.fill = make_fill(_strip_hash(rule["fill"]))

            fmt = col.get("format")
            if fmt:
                cell.number_format = fmt

    last_data_row = first_data_row + len(rows_data) - 1

    # Totals row
    if totals_spec and rows_data:
        totals_row_num = last_data_row + 1
        totals_fill_spec = totals_spec.get("fill", "primary")
        totals_fill_hex = resolve_color(totals_fill_spec, brand)
        totals_bold: bool = totals_spec.get("bold", True)
        label_col_key: str = totals_spec.get("label_column", "")
        totals_label: str = totals_spec.get("label", "TOTAL")
        sum_keys: List[str] = totals_spec.get("sum_columns", [])
        totals_fmt: Optional[str] = totals_spec.get("format")

        for ci, col in enumerate(active_columns, start=1):
            key = col.get("data_key", "")
            if key == label_col_key:
                cell_value: Any = totals_label
            elif key in sum_keys:
                total_val = sum(
                    r.get(key, 0) or 0
                    for r in rows_data
                    if isinstance(r, dict) and isinstance(r.get(key), (int, float))
                )
                cell_value = total_val
            else:
                cell_value = None

            cell = ws.cell(row=totals_row_num, column=ci, value=cell_value)
            cell.font = make_font(
                name=body_font_name,
                size=body_font_size,
                bold=totals_bold,
                color="FFFFFF",
            )
            if totals_fill_hex:
                cell.fill = make_fill(totals_fill_hex)
            cell.alignment = Alignment(vertical="center")

            if cell_value is not None and key in sum_keys:
                fmt = totals_fmt or col.get("format")
                if fmt:
                    cell.number_format = fmt

        last_data_row = totals_row_num

    # Footer / note
    if note:
        note_row = last_data_row + 2
        note_cell = ws.cell(row=note_row, column=1, value=note)
        note_cell.font = make_font(
            name=body_font_name, size=body_font_size - 1,
            italic=True, color=text_muted_hex,
        )
        note_cell.alignment = Alignment(wrap_text=True)
        if len(active_columns) > 1:
            ws.merge_cells(
                start_row=note_row, start_column=1,
                end_row=note_row, end_column=len(active_columns),
            )
        ws.row_dimensions[note_row].height = 30

    # Charts
    for chart_spec in charts_spec:
        try:
            _build_chart(
                ws, chart_spec, rows_data, active_columns, brand,
                first_data_row, last_data_row,
            )
        except Exception as exc:
            print(
                f"WARNING: chart '{chart_spec.get('id', '?')}' failed: {exc}",
                file=sys.stderr,
            )


# ---------------------------------------------------------------------------
# list sheet builder
# ---------------------------------------------------------------------------

def build_list_sheet(
    ws, sheet_spec: dict, data: dict, brand: dict, wb_style: dict
) -> None:
    """
    Sheet type 'list' — array of strings, with a row-index column and the
    sentinel data_key '_value'. Delegates to build_table_sheet after ensuring
    the items are accessible at the right path and columns are correct.
    """
    # list sheets use the same table machinery; all special handling is already
    # in build_table_sheet (_value sentinel, row_index type, header_note).
    build_table_sheet(ws, sheet_spec, data, brand, wb_style)


# ---------------------------------------------------------------------------
# Chart builder
# ---------------------------------------------------------------------------

def _build_chart(
    ws,
    chart_spec: dict,
    rows_data: List[dict],
    active_columns: List[dict],
    brand: dict,
    first_data_row: int,
    last_data_row: int,
) -> None:
    """
    Write a helper data block to the worksheet and add an openpyxl chart.

    For aggregate='sum', groups by x_axis key and sums y_axis values, then
    writes those computed values to a helper block. openpyxl charts must
    reference cell ranges — they cannot take in-memory lists directly.
    """
    chart_type: str = chart_spec.get("type", "BarChart")
    title: str = chart_spec.get("title", "")
    x_key: Optional[str] = chart_spec.get("x_axis")
    y_key: Optional[str] = chart_spec.get("y_axis")
    aggregate: Optional[str] = chart_spec.get("aggregate")
    series_color_spec = chart_spec.get("series_color", "primary")
    placement: str = chart_spec.get("placement", "right_of_table")
    width_cols: int = chart_spec.get("width_cols", 8)
    height_rows: int = chart_spec.get("height_rows", 16)
    y_label: str = chart_spec.get("y_axis_label", y_key or "")

    if not x_key or not y_key or not rows_data:
        return

    # Resolve series colors
    color1: str = resolve_color(series_color_spec, brand) or "1F3864"
    color2: str = resolve_color("highlight", brand) or "52B788"
    color3: str = resolve_color("text_muted", brand) or "6B7280"

    # Build (keys, vals) — aggregate if requested
    if aggregate == "sum":
        agg: Dict[str, float] = defaultdict(float)
        order: List[str] = []
        for row in rows_data:
            x_val = row.get(x_key, "Unknown") if isinstance(row, dict) else "Unknown"
            y_val = row.get(y_key, 0) if isinstance(row, dict) else 0
            if isinstance(y_val, (int, float)):
                if x_val not in agg:
                    order.append(str(x_val))
                agg[str(x_val)] += y_val
        chart_keys: List[Any] = order
        chart_vals: List[float] = [agg[k] for k in order]
    else:
        chart_keys = [
            row.get(x_key, "") if isinstance(row, dict) else "" for row in rows_data
        ]
        chart_vals = [
            row.get(y_key, 0) if isinstance(row, dict) else 0 for row in rows_data
        ]

    if not chart_keys:
        return

    # Write helper data block to the right of the table (cols beyond active_columns)
    helper_col = len(active_columns) + 2
    helper_start = 2  # leave row 1 for the column header

    ws.cell(row=1, column=helper_col, value=x_key)
    ws.cell(row=1, column=helper_col + 1, value=y_key)
    for i, (k, v) in enumerate(zip(chart_keys, chart_vals)):
        ws.cell(row=helper_start + i, column=helper_col, value=k)
        ws.cell(row=helper_start + i, column=helper_col + 1, value=v)

    n = len(chart_keys)
    cats_ref = Reference(
        ws, min_col=helper_col, min_row=helper_start, max_row=helper_start + n - 1,
    )
    data_ref = Reference(
        ws, min_col=helper_col + 1, min_row=1, max_row=helper_start + n - 1,
    )

    # Instantiate chart
    if chart_type == "BarChart":
        chart: Any = BarChart()
        chart.type = "col"         # vertical column chart (not horizontal bar)
        chart.grouping = "clustered"
    elif chart_type == "LineChart":
        chart = LineChart()
    elif chart_type == "PieChart":
        chart = PieChart()
    else:
        chart = BarChart()
        chart.type = "col"
        chart.grouping = "clustered"

    chart.title = title
    chart.style = 10
    if y_label and hasattr(chart, "y_axis"):
        chart.y_axis.title = y_label

    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Apply series colors
    colors = [color1, color2, color3]
    for si, ser in enumerate(chart.series):
        c = colors[si % len(colors)]
        try:
            ser.graphicalProperties.solidFill = f"FF{c}"
            ser.graphicalProperties.line.solidFill = f"FF{c}"
        except Exception:
            pass

    chart.width = width_cols * 1.8   # cm (approx)
    chart.height = height_rows * 0.5  # cm

    # Anchor
    if placement == "right_of_table":
        anchor_cell = f"{get_column_letter(helper_col + 3)}2"
    else:
        anchor_cell = chart_spec.get("anchor", f"{get_column_letter(len(active_columns) + 3)}2")

    ws.add_chart(chart, anchor_cell)


# ---------------------------------------------------------------------------
# Auto-derive sheets when xlsx.json is absent
# ---------------------------------------------------------------------------

def _auto_derive_sheets(data: dict) -> List[dict]:
    """
    Fallback: build basic sheet specs from every list key in data.
    Returns specs in insertion order.
    """
    specs = []
    for order, (key, val) in enumerate(data.items(), start=2):
        if not isinstance(val, list) or not val:
            continue
        name = sanitize_sheet_name(key.replace("_", " ").title())
        first = val[0]
        if isinstance(first, dict):
            columns = [
                {"header": k.replace("_", " ").title(), "data_key": k}
                for k in first.keys()
            ]
        else:
            columns = [
                {"header": "#", "type": "row_index", "width": 5},
                {"header": "Value", "data_key": "_value", "width": 60, "wrap": True},
            ]
        specs.append({
            "id": key,
            "name": name,
            "order": order,
            "type": "table" if isinstance(first, dict) else "list",
            "data_path": key,
            "columns": columns,
        })
    return specs


def _build_auto_summary_sheet(
    ws, data: dict, brand: dict, wb_style: dict
) -> None:
    """Summary sheet used only when no xlsx.json spec exists."""
    body_font_name = wb_style.get("body_font", "Calibri")
    body_font_size = wb_style.get("body_font_size", 9)
    header_fill_hex = resolve_color("brand.primary_color", brand)
    header_font_hex = "FFFFFF"

    for ci, hdr in enumerate(["Section", "Key Metric", "Value"], start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = make_font(name=body_font_name, size=body_font_size, bold=True, color=header_font_hex)
        if header_fill_hex:
            cell.fill = make_fill(header_fill_hex)

    row = 2
    for key, val in data.items():
        if isinstance(val, (list, dict)):
            continue
        ws.cell(row=row, column=1, value=key.replace("_", " ").title())
        ws.cell(row=row, column=2, value="value")
        ws.cell(row=row, column=3, value=str(val) if val is not None else "")
        row += 1

    for ci in range(1, 4):
        col_letter = get_column_letter(ci)
        max_len = max(
            (len(str(ws.cell(row=r, column=ci).value or ""))
             for r in range(1, row)),
            default=10,
        )
        ws.column_dimensions[col_letter].width = min(50, max(10, max_len + 2))


# ---------------------------------------------------------------------------
# Workbook-level styles
# ---------------------------------------------------------------------------

def _apply_workbook_style(wb: Workbook, wb_style: dict, brand: dict) -> None:
    """
    Apply per-sheet tab colors and print settings from workbook_style.

    Tab color: openpyxl tabColor is a bare 6-char hex (no '#', no alpha).
    We write it as 6 hex chars derived from resolve_color().
    """
    tab_colors: dict = wb_style.get("tab_colors", {})
    print_settings: dict = wb_style.get("print_settings", {})

    for ws in wb.worksheets:
        # Tab color
        raw_tc = tab_colors.get(ws.title)
        if raw_tc:
            tc_hex6 = resolve_color(raw_tc, brand)
            if tc_hex6:
                ws.sheet_properties.tabColor = tc_hex6

        # Print settings
        if print_settings.get("orientation") == "landscape":
            ws.page_setup.orientation = "landscape"
        if print_settings.get("fit_to_page"):
            ws.page_setup.fitToPage = True
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
        if print_settings.get("repeat_header_rows"):
            ws.print_title_rows = "1:1"


# ---------------------------------------------------------------------------
# Main workbook builder
# ---------------------------------------------------------------------------

def build_workbook(
    template: str,
    data: dict,
    brand: dict,
    output_path: str,
    templates_dir: str,
) -> None:
    """
    Build and save the Excel workbook.

    1. Ensure decarb_plan_total is computed.
    2. Load xlsx.json if present; otherwise auto-derive.
    3. Build each sheet in order, skipping empty omit_if_empty sheets.
    4. Apply workbook-level style (tab colors, print settings).
    5. Save to output_path.
    """
    ensure_decarb_plan_total(data)

    xlsx_spec_path = os.path.join(templates_dir, template, "xlsx.json")
    if os.path.exists(xlsx_spec_path):
        with open(xlsx_spec_path, encoding="utf-8") as fh:
            xlsx_spec: dict = json.load(fh)
        sheets_spec = sorted(
            xlsx_spec.get("sheets", []), key=lambda s: s.get("order", 999)
        )
        use_auto_summary = False
    else:
        xlsx_spec = {}
        sheets_spec = _auto_derive_sheets(data)
        use_auto_summary = True

    wb_style: dict = xlsx_spec.get("workbook_style", {})

    wb = Workbook()
    # Remove the default empty sheet Excel always creates
    if wb.worksheets:
        wb.remove(wb.active)

    # Auto-summary (fallback only — when xlsx.json already defines a summary, skip)
    if use_auto_summary:
        ws_sum = wb.create_sheet("Summary")
        _build_auto_summary_sheet(ws_sum, data, brand, wb_style)

    for sheet_spec in sheets_spec:
        sheet_name = sanitize_sheet_name(sheet_spec.get("name", "Sheet"))
        sheet_type: str = sheet_spec.get("type", "table")

        # Evaluate omit_if_empty at the workbook level (before creating the sheet)
        if sheet_spec.get("omit_if_empty"):
            dotpath = sheet_spec.get("data_path", "")
            candidates: List[Any] = resolve_path(data, dotpath) or []
            filter_spec = sheet_spec.get("filter")
            if filter_spec:
                field = filter_spec.get("field")
                if filter_spec.get("not_null") and field:
                    candidates = [
                        r for r in candidates
                        if isinstance(r, dict) and r.get(field) not in (None, "", [])
                    ]
            if not candidates:
                continue

        ws = wb.create_sheet(title=sheet_name)

        if sheet_type == "key_value":
            build_key_value_sheet(ws, sheet_spec, data, brand, wb_style)
        elif sheet_type == "table":
            build_table_sheet(ws, sheet_spec, data, brand, wb_style)
        elif sheet_type == "list":
            build_list_sheet(ws, sheet_spec, data, brand, wb_style)
        else:
            print(
                f"WARNING: unknown sheet type '{sheet_type}' for '{sheet_name}'",
                file=sys.stderr,
            )

    _apply_workbook_style(wb, wb_style, brand)

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _coerce_json(value: Any) -> Any:
    """Accept either a pre-parsed dict/list or a JSON string; always return parsed."""
    if isinstance(value, str):
        return json.loads(value)
    return value


def _load_json_file(path: str, label: str) -> dict:
    try:
        with open(path, encoding="utf-8") as fh:
            return json.load(fh)
    except FileNotFoundError:
        print(f"ERROR: {label} not found: {path}", file=sys.stderr)
        sys.exit(1)
    except json.JSONDecodeError as exc:
        print(f"ERROR: {label} is not valid JSON: {exc}", file=sys.stderr)
        sys.exit(1)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build a branded .xlsx companion for a soapbox-report-skill report.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument(
        "--config",
        help="JSON config file containing template, data, brand, output, templates_dir.",
    )
    parser.add_argument("--template", help="Template name (e.g. 'rsra').")
    parser.add_argument(
        "--data",
        help="Report data as a JSON string (or path to JSON file if prefixed with @).",
    )
    parser.add_argument(
        "--brand",
        help="Brand config as a JSON string (or path to JSON file if prefixed with @).",
    )
    parser.add_argument("--output", help="Output path for the .xlsx file.")
    parser.add_argument(
        "--templates-dir",
        dest="templates_dir",
        default=os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "templates"),
        help="Path to the templates directory (default: ../templates relative to this script).",
    )

    args = parser.parse_args()

    # Load from --config if provided; CLI flags override
    if args.config:
        cfg = _load_json_file(args.config, "--config")
    else:
        cfg = {}

    template: Optional[str] = args.template or cfg.get("template")
    output: Optional[str] = args.output or cfg.get("output")
    templates_dir: str = args.templates_dir or cfg.get("templates_dir", args.templates_dir)

    # data / brand: accept JSON string, pre-parsed object, or @filepath notation
    raw_data = args.data or cfg.get("data", "{}")
    raw_brand = args.brand or cfg.get("brand", "{}")

    def _resolve(raw: Any) -> Any:
        if isinstance(raw, str) and raw.startswith("@"):
            return _load_json_file(raw[1:], "data/brand file")
        return _coerce_json(raw)

    data = _resolve(raw_data)
    brand = _resolve(raw_brand)

    # Validate required args
    errors: List[str] = []
    if not template:
        errors.append("--template is required")
    if not output:
        errors.append("--output is required")
    if errors:
        for msg in errors:
            print(f"ERROR: {msg}", file=sys.stderr)
        parser.print_usage(sys.stderr)
        sys.exit(1)

    try:
        build_workbook(
            template=template,
            data=data,
            brand=brand,
            output_path=output,
            templates_dir=os.path.abspath(templates_dir),
        )
    except Exception as exc:
        import traceback
        print(f"ERROR: workbook build failed: {exc}", file=sys.stderr)
        traceback.print_exc(file=sys.stderr)
        sys.exit(1)

    print(output)


if __name__ == "__main__":
    main()
